import os
import logging
import json
from collections import OrderedDict

import boto3
import datetime
from urllib.request import urlretrieve
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.forms.models import model_to_dict
from django.http.response import HttpResponse

from .models import CompanyGroup, Company, Car
from .forms import UploadVehicleFileForm


logger = logging.getLogger(__name__)


def get_s3_client():
    aws_access_key_id = os.environ.get('AWS_ACCESS_KEY_ID')
    aws_secret_access_key = os.environ.get('AWS_SECRET_ACCESS_KEY')
    return boto3.client(
        's3',
        aws_access_key_id=aws_access_key_id,
        aws_secret_access_key=aws_secret_access_key,
    )


def upload_image_from_url(url, id=None):
    id = str(id) if id is not None else ''
    tail, head = url.rsplit('/', 1) if '/' in url else ('', '')
    filename = '{}_{}'.format(id, head)
    try:
        urlretrieve(url, filename)
    except Exception as e:
        logger.warning('error: urlretrieve: %s: %s', type(e).__name__, str(e))
        return
    bucket_name = 'pictures.anno1453'

    try:
        s3 = get_s3_client()
        url = 'https://s3.amazonaws.com/{}/{}'.format(bucket_name, filename)
        result = s3.upload_file(filename, bucket_name, filename, ExtraArgs={'ACL': 'public-read'})
        logger.warning('upload_file result %s', result, )

    except Exception as e:
        logger.error('error: upload_file: %s: %s', str(type(e)), str(e))
        logger.warning('upload_file: args: %s %s %s', filename, bucket_name, filename)

    else:
        return url

    finally:
        if os.path.exists(filename):
            os.remove(filename)
        pass


def index(request):
    return HttpResponseRedirect('/cars/')


def handle_car(raw_car_data):
    company = Company.objects.filter(inn=raw_car_data.get('CompanyINN')).first()
    logger.warning('COMPANY %s %s %s', company, raw_car_data.get('CompanyINN'), Company.objects.all())

    data = {}
    data['company'] = company
    data['vin'] = raw_car_data['VIN']
    data['model'] = raw_car_data['Model']
    data['year'] = int(raw_car_data['Year'])
    data['price'] = float(raw_car_data['Price'])
    data['gosnomer'] = raw_car_data['Gosnomer']
    data['used'] = float(raw_car_data.get("Kilometrage", '0')) >= 10000.0
    data['forloan'] = raw_car_data.get("ForLoan") == "1"

    car = Car.objects.create(**data)
    for car_image_url in [car_image.get('-url') for car_image in raw_car_data.get("Images", {}).get("CarImage", [])]:
        logger.warning('car_image_url %s', car_image_url)
        image_url = upload_image_from_url(car_image_url, id=car.id)
        if image_url:
            car.images.create(url=image_url)
    car.save()
    logger.warning('CAR %s', car.id)


def handle_cars(raw_cars_data):
    for raw_company_group_data in raw_cars_data.get('Cars', {}).get('Companies', []):
        i_d = raw_company_group_data.get('CompanyGroupId')
        company_group_data = {}
        company_group_data['i_d'] = i_d
        company_group_data['name'] = raw_company_group_data.get('CompanyGroupName')
        company_group = CompanyGroup.objects.filter(i_d=i_d).first()
        if not company_group:
            company_group = CompanyGroup.objects.create(**company_group_data)
        for raw_company_data in raw_company_group_data.get('GroupCompanies', []):
            inn = raw_company_data.get('CompanyINN')

            company = Company.objects.filter(inn=inn).first()
            if not company:
                company_data = {}
                company_data['inn'] = inn
                company_data['group'] = company_group
                company_data['name'] = raw_company_data.get('CompanyName')
                Company.objects.create(**company_data)

    for car in raw_cars_data.get('Cars', {}).get('Car', []):
        handle_car(car)


def handle_uploaded_file(f):
    try:
        content = f.read()
        raw_data = json.loads(content)
        handle_cars(raw_data)

    except Exception as e:
        err = '{}: {}'.format(type(e).__name__, str(e))
        import traceback
        logger.error('Error: %s %s', err, traceback.format_exc())
        return err


xlsx_field_map = OrderedDict([
    ('company', 'наименование компании'),
    ('inn', 'инн компании'),
    ('vin', 'VIN'),
    ('model', 'модель'),
    ('year', 'год'),
    ('price', 'цена'),
    ('gosnomer', 'госномер'),
    ('unused', 'новая'),
    ('forloan', 'для продажи'),
])


def make_xlsx_response(ids, field_map=None):
    try:
        ids = list(ids)

    except Exception as e:
        ids = None
        logger.warning('make_xlsx_response %s %s', type(e).__name__, str(e))

    logger.warning('make_xlsx_response(%s)', ids)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cars.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Автомобили"

    row_num = 0

    columns = [(value, 20) for value in field_map.values()]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for car in Car.objects.all():
        if ids and car.id not in ids:
            continue

        row_num += 1
        car_data = car_to_dict(car, field_map.keys())
        row = [str(car_data.get(key)) for key in field_map]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response


def car_list(request):
    format = request.GET.get('format', 'html')

    if format == 'xlsx':
        return make_xlsx_response(map(int, request.GET.getlist('id')), xlsx_field_map)

    return render(request, 'car_list.html', {'cars': Car.objects.all()})


def cars_del(request):
    ids = list(map(int, request.GET.getlist('id')))
    for id in ids:
        Car.objects.get(id=id).delete()
    return HttpResponseRedirect('/cars/')


def cars_json(request):
    cars_data = [car_to_dict(car) for car in Car.objects.all()]

    return JsonResponse(cars_data, safe=False)


def car_to_dict(car, fields=None):
    fields = fields if fields is not None else []
    car_data = model_to_dict(car, fields)
    car_data['created'] = datetime.date.strftime(car.created, '%m/%d/%Y') if car.created else ''
    car_data['imagesurl'] = car.imagesurl()
    car_data['company'] = car.company.name if car.company else ''
    car_data['inn'] = car.company.inn if car.company else ''
    car_data['unused'] = 'Да' if not car.used else 'Нет'
    car_data['forloan'] = 'Да' if car.forloan else 'Нет'
    return car_data


def car_add(request):
    logger.warning('car_add %s %s %s', request.method, request.GET, request.POST)
    if request.method == 'POST':
        form = UploadVehicleFileForm(request.POST, request.FILES)
        if form.is_valid():
            error = handle_uploaded_file(request.FILES['file'])
            if error is not None:
                return render(request, 'hello/car_add_error.html', {'error': error})
            return HttpResponseRedirect('/cars/')
    else:
        form = UploadVehicleFileForm()
    return render(request, 'hello/car_add.html', {'form': form})


def car_detail(request, pk):
    logger.warning('car_detail %s', pk)
    car = Car.objects.get(pk=pk)
    return render(request, 'hello/car_detail.html', {'id': car.id, 'vin': car.vin, 'images': car.images.all()})
